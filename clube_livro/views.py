from django.shortcuts import render, redirect, get_object_or_404
from .models import Livro, WishList
from .forms import RegistrarLivro, RegistraLista

#Filtro
from .filters import LivrosFilter
#Excel sheet
from openpyxl import Workbook
from django.http import HttpResponse
from datetime import datetime
from datetime import timedelta

from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

 
def mainView(request):
    return render( request, 'main.html',)



def profileView(request):
    quantidade = len(Livro.objects.filter(auth=request.user))
    return render( request, 'profile.html', {'quantidade':quantidade})


def addView(request):
    form = RegistrarLivro(request.POST, request.FILES, )
    if form.is_valid():
        #o obj guarda o save do form, que não eh executado no banco de dados por causa do commit=false
        obj = form.save(commit=False)
        obj.auth = request.user
        obj.save()
        return redirect("read")
    return render(request, 'add.html', {'form':form, })


def readView(request):
    
    #Pega apenas os objetos atrelados à aquele user
    livros = Livro.objects.filter(auth=request.user)

    myFilter = LivrosFilter(request.GET, queryset=livros)
    livros = myFilter.qs

    return render(request, 'read.html', {'livros':livros, 'myFilter': myFilter, })


def detailView(request, id):
    livro = get_object_or_404(Livro, pk=id)
    return render(request, 'detail.html', {'livro':livro})


def updateView(request, id):
    name = get_object_or_404(Livro, pk=id)
    form = RegistrarLivro(request.POST or None, request.FILES or None, instance=name)
    if form.is_valid():
        form.save()
        return redirect("read")
    return render(request, 'update.html', {'form': form})



def deleteView(request, id):
    book = get_object_or_404(Livro, pk=id)
    if request.method == 'POST':
        book.delete()
        return redirect('read')
    return render(request, 'delete.html', {'book': book})



def authView(request):
    return render(request, 'author.html', {})

#views Lista de desejo

def wish_list(request):
    desejos = WishList.objects.filter(list_auth=request.user)
    return render(request, 'wish_list/list.html', {'desejos':desejos})


def add_wish_list(request):
    form = RegistraLista(request.POST)
    if form.is_valid():
        obj = form.save(commit=False)
        obj.list_auth = request.user
        obj.save()
        return redirect('wish')
    return render(request, 'wish_list/add_list.html', {'form':form})


#função para baixar a database em excel

def excel_books(request):
    
    livros_lista = Livro.objects.all()

    #estudar depois o que isso aqui faz
    response = HttpResponse(
        content_type='application/ms-excel',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-livros.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )

    workbook = Workbook()

    #Get active worksheet/tab:
    worksheet = workbook.active
    worksheet.title = "Livros"

    #Definindo as colunas do documento
    columns = ['auth','book_title','book_type','book_sinopse','user_rating']
    row_num = 1

    #This atribue value to the columns headers
    for col_number, column_title in enumerate(columns,1):
        cell = worksheet.cell(row = row_num, column=col_number)
        cell.value = column_title

    #Define the data for each cell in the row
    for livro in livros_lista:
        row_num +=1

        row=[
            str(livro.auth),
            livro.book_title,
            livro.book_type,
            livro.book_sinopse,
            livro.user_rating,
        ]

        for col_number, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_number)
            cell.value = cell_value
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    #for fix the width of th column
    dimensions = [40, 25, 14, 14, 11]
    for col_num, width in enumerate(dimensions, 1):
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = width


    workbook.save(response)

    return response
        







